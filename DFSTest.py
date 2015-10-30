#https://automatetheboringstuff.com
import openpyxl   #http://openpyxl.readthedocs.org
import xlsxwriter   #https://pypi.python.org/pypi/XlsxWriter / http://xlsxwriter.readthedocs.org
#hex color codes http://cloford.com/resources/colours/500col.htm

# http://www.footballoutsiders.com/stats/teamdef  Defensive Ranking website


defRowNum = 2   #column  of DFS spreadsheet
rowNum = 2   #row number of DFS spreadsheet

fdWB = openpyxl.load_workbook('FanDuelPlayerList.xlsx')   #read FD player stat xlsx file
fdSheet = fdWB.get_active_sheet()

defenseWB = openpyxl.load_workbook('DefensiveRanking.xlsx')    #read defesive ranking xlsx file
defenseSheet = defenseWB.get_active_sheet()

dfsWB = xlsxwriter.Workbook('DFStats.xlsx')   #create xlsx file
dfsSheet = dfsWB.add_worksheet('DFSTest')

bold = dfsWB.add_format({'bold': True})
redCell = dfsWB.add_format({'bg_color': '#CD5C5C'})  #indianred #CD5C5C	hex code
greenCell = dfsWB.add_format({'bg_color': '#00FF7F'})   #springgreen #00FF7F hex code
blueCell = dfsWB.add_format({'bg_color': '#00BFFF'})   #deepskyblue #00BFFF hex code

dfsSheet.write(0, 0, 'Player Team', bold)
dfsSheet.write(0, 1, 'Position', bold)
dfsSheet.write(0, 2, 'First Name', bold)
dfsSheet.write(0, 3, 'Last Name', bold)
dfsSheet.write(0, 4, 'Opponent', bold)
dfsSheet.write(0, 5, 'Opp. D. Pass Rank', bold)
dfsSheet.write(0, 6, 'Opp. D. Rush Rank', bold)
dfsSheet.write(0, 7, 'FPPG', bold)
dfsSheet.write(0, 8, 'FD Salary', bold)
dfsSheet.write(0, 9, 'Projected Value', bold)
dfsSheet.write(0, 10, 'Injury Status', bold)
dfsSheet.write(0, 11, 'Injury Details', bold)

while rowNum <= fdSheet.get_highest_row():
    defRowNum = 2
    playerTeam = fdSheet['I' + str(rowNum)].value
    playerOpp = fdSheet['J' + str(rowNum)].value
    oppTeamName = defenseSheet['A' + str(defRowNum)].value
    oppPassRank = defenseSheet['G'+ str(defRowNum)].value
    oppRushRank = defenseSheet['I'+ str(defRowNum)].value
    playerPosition = fdSheet['B' + str(rowNum)].value
    playerFirstName = fdSheet['C' + str(rowNum)].value
    playerLastName = fdSheet['D' + str(rowNum)].value
    injuryStatus = fdSheet['K' + str(rowNum)].value
    injuryDetails = fdSheet['L' + str(rowNum)].value
    fdFPPG = fdSheet['E' + str(rowNum)]   #FD fanatsy points per game
    fdSalary = fdSheet['G' + str(rowNum)]   #FD salary amount
    projectedValue = ((fdFPPG.value / fdSalary.value) * 1000)   #FD projected value based on FPPG / Salary

    dfsSheet.write(rowNum-1, 0, playerTeam)   #write all relevant data to xlsx sheet
    dfsSheet.write(rowNum-1, 1, str(playerPosition))
    dfsSheet.write(rowNum-1, 2, str(playerFirstName))
    dfsSheet.write(rowNum-1, 3, str(playerLastName))
    dfsSheet.write(rowNum-1, 4, playerOpp)
    
    while defRowNum <= defenseSheet.get_highest_row():
        if playerOpp != oppTeamName:
            defRowNum += 1
        else:
            dfsSheet.write(rowNum-1, 5, oppPassRank)
            dfsSheet.write(rowNum-1, 6, oppRushRank)
        
    dfsSheet.write(rowNum-1, 7, str(fdFPPG.value))
    dfsSheet.write(rowNum-1, 8, str(fdSalary.value))

    if projectedValue > 2.25:
        dfsSheet.write(rowNum-1, 9, projectedValue, greenCell)
    elif projectedValue > 1.75:
        dfsSheet.write(rowNum-1, 9, projectedValue, blueCell)
    elif projectedValue < 1.25:
        dfsSheet.write(rowNum-1, 9, projectedValue, redCell)
    else:
        dfsSheet.write(rowNum-1, 9, projectedValue)

    dfsSheet.write(rowNum-1, 10, injuryStatus)
    dfsSheet.write(rowNum-1, 11, injuryDetails)
    rowNum += 1

dfsWB.close()

#OLD CODE, HOLDING JUST IN CASE
"""
import webbrowser
import requests
import linecache   #https://docs.python.org/2/library/linecache.html
import bs4
from selenium import webdriver
lineNum = 594  #line number of where player stats starts
lineEndNum = 969   #line where the player stats end
resSite = requests.get('http://rotoguru1.com/cgi-bin/fyday.pl?week=5&game=fd&scsv=1')   #open/download website text info
dfFile = open('DFStats.txt', 'wb')   #open text file in binary write mode
for chunk in resSite.iter_content(10000):   #write text info to file
    dfFile.write(chunk)
dfFile.close()
while lineNum <= lineEndNum:
    testLine = linecache.getline('DFStats.txt', lineNum).replace(';', ' ').split()  #get starting line for player stats and store in list, replace ; and then split the white space.
    print(testLine)
    lineNum += 1
    for i in range(len(testLine)):
        sheet.write(rowNum, colNum, testLine[i])   #write data from testLine into correct column
        colNum += 1
    colNum = 0   # reset colNum on newline
    rowNum +=1
"""
